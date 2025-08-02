from openpyxl import load_workbook
import pandas as pd

def process_file(file_path, file_type):
    if file_type == 'fluid' or file_type == 'cement':
        wb = load_workbook(file_path, data_only=True) # Load the workbook
        ws = wb['Signatures that need to be read']  # Specifying the sheet
        ws_signature = wb['Signatures that need to be read']  # Specifying the sheet
        ws_price = wb['price and cost sheet']  # Specifying the sheet

    if file_type == 'fluid':
        # Get Client
        client = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "client":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    client = next_cell.value
                    break
            if client is not None:
                break

        # Get Well
        well = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "well":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    well = next_cell.value
                    break
            if well is not None:
                break

        # get basic
        basic = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "basic":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    basic = next_cell.value
                    break
            if basic is not None:
                break

        # get system
        system = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "system":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 2)
                    system = next_cell.value
                    break
            if system is not None:
                break

        # get income

        # get material income
        # Load price data with correct headers
        price_data = ws_price.values
        next(price_data)  # Skip row 1 (empty)
        headers = next(price_data)  # Row 2: 'Description', 'price', 'unit', 'cost'
        df_price = pd.DataFrame(price_data, columns=headers)
        df_price.dropna(subset=['Description', 'price'], inplace=True)

        # Material section: starts at row 10, ends at 37
        total_material_income = 0.0

        for row in range(10, 38):
            material_name = ws_signature.cell(row=row, column=2).value  # B
            quantity = ws_signature.cell(row=row, column=3).value       # C

            if material_name and quantity:
                match = df_price[df_price['Description'] == material_name]
                if not match.empty:
                    price = match.iloc[0]['price']
                    income = quantity * price
                    total_material_income += income

        # get personnel income
        # Calculate Personnel Income
        total_personnel_income = 0.0

        for row in range(44, 52):  # Row 44–51 in Excel (personnel section)
            personnel_name = ws_signature.cell(row=row, column=6).value  # F column
            days = ws_signature.cell(row=row, column=9).value            # I column

            if personnel_name and days:
                match = df_price[df_price['Description'] == personnel_name]
                if not match.empty:
                    price_per_day = match.iloc[0]['price']
                    income = days * price_per_day
                    total_personnel_income += income

        # get other items income
        # Calculate income from other items
        total_other_items_income = 0.0

        for row in range(31, 40):  # Scan rows 31 to 39
            item_name = ws_signature.cell(row=row, column=6).value  # Column F = project name
            quantity = ws_signature.cell(row=row, column=7).value   # Column G = quantity

            if item_name and quantity:
                match = df_price[df_price['Description'] == item_name]
                if not match.empty:
                    price = match.iloc[0]['price']
                    income = quantity * price
                    total_other_items_income += income

        # get onshore services income
        # Calculate income from onshore services
        total_onshore_services_income = 0.0

        for row in range(44, 52):  # Rows 44–51 inclusive
            service_name = ws_signature.cell(row=row, column=2).value  # Column B
            quantity = ws_signature.cell(row=row, column=3).value      # Column C

            if service_name and quantity:
                match = df_price[df_price['Description'] == service_name]
                if not match.empty:
                    price = match.iloc[0]['price']
                    income = quantity * price
                    total_onshore_services_income += income

        # get equipment income
        # Equipment income calculation
        total_equipment_income = 0.0

        for row in range(10, 29):  # Rows 10 to 28
            equipment_name = ws_signature.cell(row=row, column=6).value  # Column F
            days = ws_signature.cell(row=row, column=9).value            # Column I

            if equipment_name and days:
                match = df_price[df_price['Description'] == equipment_name]
                if not match.empty:
                    price = match.iloc[0]['price']
                    income = days * price
                    total_equipment_income += income

        # get system income
        total_system_income = 0.0

        # Rows 40–42 (just below "system" label)
        for row in range(40, 43):
            name = ws_signature.cell(row=row, column=1).value  # Column A
            qty = ws_signature.cell(row=row, column=3).value  # Column C

            if name and qty:
                match = df_price[df_price['Description'].astype(str).str.strip() == str(name).strip()]
                if not match.empty:
                    price = match.iloc[0]['price']
                    total_system_income += qty * price

        total_income = (total_material_income + total_personnel_income + total_other_items_income 
        + total_onshore_services_income + total_equipment_income + total_system_income)

        # get well type
        well_type = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "well type":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    well_type = next_cell.value
                    break
            if well_type is not None:
                break

        # get manager
        manager = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "manager":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    manager = next_cell.value
                    break
            if manager is not None:
                break

        # get area
        area = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "area":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    area = next_cell.value
                    break
            if area is not None:
                break

        # get start time
        # Search for the cell with "使用时间"
        start_time = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value == "使用时间":
                    # Get the value one row below (same column)
                    start_time = ws.cell(row=cell.row + 1, column=cell.column).value.date()
                    break
            if start_time:
                break

        # get end time
        # Search for "使用时间"
        end_time = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value == "使用时间":
                    # One row below, next column (column + 1)
                    end_time = ws.cell(row=cell.row + 1, column=cell.column + 1).value.date()
                    break
            if end_time:
                break

        # get well class
        well_class = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "well class":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    well_class = next_cell.value
                    break
            if well_class is not None:
                break

        # get important type
        important_type = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "important type":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    important_type = next_cell.value
                    break
            if important_type is not None:
                break

        # get risk
        risk = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "risk":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    risk = next_cell.value
                    break
            if risk is not None:
                break

        # get deep
        deep = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "deep":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    deep = next_cell.value
                    break
            if deep is not None:
                break

        # get tvd
        tvd = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "tvd":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    tvd = next_cell.value
                    break
            if tvd is not None:
                break

        # get cost
        # cost per material
        # Prepare list to store cost records
        cost_records = []

        # Read materials from Signature sheet rows 10–37
        for row in range(10, 38):
            material_name = ws_signature.cell(row=row, column=2).value  # Column B
            quantity = ws_signature.cell(row=row, column=3).value       # Column C

            if material_name and quantity:
                match = df_price[df_price['Description'] == material_name]
                if not match.empty:
                    unit_cost = match.iloc[0]['cost']
                    total_cost = quantity * unit_cost

                    record = {
                        'client': client,
                        'well': well,
                        'basic': basic,
                        'system': system,
                        'materialname': material_name,
                        'use': quantity,
                        'cost': unit_cost,
                        'cost * use': total_cost
                    }
                    cost_records.append(record)

        # Output result as a DataFrame for easier viewing/export
        df_cost_records = pd.DataFrame(cost_records)

        # total cost
        # Initialize total cost
        total_cost = 0.0

        # 1. Material Cost (rows 10–37)
        for row in range(10, 38):
            name = ws.cell(row=row, column=2).value
            qty = ws.cell(row=row, column=3).value
            if name and qty:
                match = df_price[df_price['Description'] == name]
                if not match.empty:
                    total_cost += qty * match.iloc[0]['cost']

        # 2. Personnel Cost (rows 44–51)
        for row in range(44, 52):
            name = ws.cell(row=row, column=6).value
            days = ws.cell(row=row, column=9).value
            if name and days:
                match = df_price[df_price['Description'] == name]
                if not match.empty:
                    total_cost += days * match.iloc[0]['cost']

        # 3. Equipment Cost (rows 10–28)
        for row in range(10, 29):
            name = ws.cell(row=row, column=6).value
            days = ws.cell(row=row, column=9).value
            if name and days:
                match = df_price[df_price['Description'] == name]
                if not match.empty:
                    total_cost += days * match.iloc[0]['cost']

        # 4. Other Items Cost (rows 31–39)
        for row in range(31, 40):
            name = ws.cell(row=row, column=6).value
            qty = ws.cell(row=row, column=7).value
            if name and qty:
                match = df_price[df_price['Description'] == name]
                if not match.empty:
                    total_cost += qty * match.iloc[0]['cost']

        # 5. Onshore Services Cost (rows 44–51)
        for row in range(44, 52):
            name = ws.cell(row=row, column=2).value
            qty = ws.cell(row=row, column=3).value
            if name and qty:
                match = df_price[df_price['Description'] == name]
                if not match.empty:
                    total_cost += qty * match.iloc[0]['cost']

        # 6. System Cost (rows 40–42)
        for row in range(40, 43):
            name = ws.cell(row=row, column=1).value  # Column A (leftmost, due to merged cells)
            qty = ws.cell(row=row, column=3).value   # Column C
            if name and qty:
                match = df_price[df_price['Description'].astype(str).str.strip() == str(name).strip()]
                if not match.empty:
                    total_cost += qty * match.iloc[0]['cost']

        # Generating report
        # Revenue records
        revenue_records = {
            'client': [client],
            'well': [well],
            'basic': [basic],
            'system': [system],
            'income': [total_income],
            'well type': [well_type],
            'manager': [manager],
            'area': [area],
            'start time': [start_time],
            'end time': [end_time],
            'well class': [well_class],
            'important type': [important_type],
            'risk': [risk],
            'deep': [deep],
            'tvd': [tvd]
        }

        df_revenue_records = pd.DataFrame(revenue_records)

        # Cost records

        # Profit list
        profit = total_income - total_cost
        profit_rate = (profit/total_income)*100
        profit_list = {
            'client': [client],
            'well': [well],
            'basic': [basic],
            'system': [system],
            'income': [total_income],
            'cost': [total_cost],
            'profit': [profit],
            'profit rate': [profit_rate]
        }

        df_profit_list = pd.DataFrame(profit_list)

    elif file_type == 'cement':
        # Get Client
        client = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "client":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    client = next_cell.value
                    break
            if client is not None:
                break

        # Get Well
        well = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "well":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    well = next_cell.value
                    break
            if well is not None:
                break

        # get basic
        basic = None
        for row in ws.iter_rows():
            for cell in row:
                if str(cell.value).strip() == "basic":
                    row_index = cell.row
                    col_index = cell.column
                    # Try value from the next column (H), and next row (row below)
                    basic = ws.cell(row=row_index + 1, column=col_index + 1).value
                    break
            if basic is not None:
                break

        # get system
        system = None
        for row in ws.iter_rows():
            for cell in row:
                if str(cell.value).strip() == "system":
                    row_index = cell.row
                    col_index = cell.column
                    # Try value from the next column (H), and next row (row below)
                    system = ws.cell(row=row_index + 1, column=col_index + 1).value
                    break
            if system is not None:
                break

        # get well type
        well_type = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "well type":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    well_type = next_cell.value
                    break
            if well_type is not None:
                break

        # get manager
        manager = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "manager":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    manager = next_cell.value
                    break
            if manager is not None:
                break

        # get area
        area = 'Binhai'
        # get start time
        start_time = '2025-6-5'

        # get end time
        end_time = '2025-7-2'

        # get well class
        well_class = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "well class":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    well_class = next_cell.value
                    break
            if well_class is not None:
                break

        # get important type
        important_type = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "important type":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    important_type = next_cell.value
                    break
            if important_type is not None:
                break

        # get risk
        risk = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "risk":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    risk = next_cell.value
                    break
            if risk is not None:
                break

        # get deep
        deep = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "deep":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    deep = next_cell.value
                    break
            if deep is not None:
                break

        # get tvd
        tvd = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "tvd":
                    # Get the value in the next column (same row)
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    tvd = next_cell.value
                    break
            if tvd is not None:
                break

        # get income

        # get cementing services income
        # --- Step 1: Build cementing price lookup dictionary ---
        cementing_prices = {}

        for row in ws_price.iter_rows(min_row=2, max_col=7):
            desc = row[1].value or row[2].value  # Column B or C
            price = row[4].value                 # Column E

            if isinstance(price, (int, float)) and desc:
                cementing_prices[str(desc).strip()] = float(price)

        # --- Step 2: Calculate total cementing income ---
        total_cementing_income = 0.0

        for row in range(33, 60):  # Adjust row range if needed
            service_name = ws_signature.cell(row=row, column=6).value  # Column F
            quantity = ws_signature.cell(row=row, column=9).value      # Column I

            if service_name and isinstance(quantity, (int, float)):
                desc = str(service_name).strip()
                price = cementing_prices.get(desc)

                if price is not None:
                    total_cementing_income += quantity * price

        # get equipment services income
        # Step 1: Build lookup dictionary from equipment tools section in price sheet
        tool_service_prices = {}

        for row in ws_price.iter_rows(min_row=16, max_row=36, max_col=7):  # rows 16–36
            desc = row[1].value or row[2].value  # Column B or C
            price = row[4].value                 # Column E

            if isinstance(price, (int, float)) and desc:
                tool_service_prices[str(desc).strip()] = float(price)

        # Step 2: Calculate total equipment tools income
        total_tools_income = 0.0

        for row in range(33, 43):  # Adjust based on actual number of rows
            desc = ws_signature.cell(row=row, column=2).value  # Column B (Description)
            quantity = ws_signature.cell(row=row, column=3).value  # Column C (Service Quantity)

            if desc and isinstance(quantity, (int, float)):
                desc = str(desc).strip()
                price = tool_service_prices.get(desc)

                if price is not None:
                    total_tools_income += quantity * price

        # get other technical services income
        # Step 1: Build price dictionary from 'other technical services' section
        other_service_prices = {}

        for row in ws_price.iter_rows(min_row=51, max_row=107, max_col=7):  # C to E
            desc = row[2].value  # Column C
            price = row[4].value  # Column E

            if desc and isinstance(price, (int, float)):
                other_service_prices[str(desc).strip()] = float(price)

        # Step 2: Calculate income from 'other technical services'
        total_other_services_income = 0.0

        for row in range(11, 32):  # Signature sheet rows
            desc = ws_signature.cell(row=row, column=2).value  # Column B
            quantity = ws_signature.cell(row=row, column=3).value  # Column C

            if desc and isinstance(quantity, (int, float)):
                desc = str(desc).strip()
                price = other_service_prices.get(desc)
                if price is not None:
                    total_other_services_income += price * quantity

        # get casing accessories income
        # Step 1: Extract casing accessory prices from the "price and cost sheet"
        casing_price_list = []

        # Adjust rows based on your images (from row 111 to 209)
        for row in ws_price.iter_rows(min_row=111, max_row=209):
            description = row[2].value  # Column C
            price = row[4].value        # Column E

            if description and isinstance(price, (int, float)):
                casing_price_list.append({
                    "Description": str(description).strip(),
                    "price": float(price)
                })

        df_casing_prices = pd.DataFrame(casing_price_list)

        # Step 2: Read descriptions and consumption values from the casing table
        casing_income = 0.0

        # Based on your sheet, casing section starts at row 10 and ends at row 31
        for row in range(10, 32):
            desc = ws_signature.cell(row=row, column=6).value  # Column F: Description
            qty = ws_signature.cell(row=row, column=8).value   # Column H: Consumption

            if desc and isinstance(qty, (int, float)):
                match = df_casing_prices[df_casing_prices["Description"] == str(desc).strip()]
                if not match.empty:
                    unit_price = match.iloc[0]["price"]
                    casing_income += qty * unit_price

        total_income = (total_cementing_income + total_tools_income + total_other_services_income + casing_income)

        # get cost
        cost_records = []

        # Step 1: Cementing Services Cost
        cementing_costs = {}
        for row in ws_price.iter_rows(min_row=2, max_col=8):
            desc = row[1].value or row[2].value  # Column B or C
            cost = row[6].value                  # Column G
            if desc and isinstance(cost, (int, float)):
                cementing_costs[str(desc).strip()] = float(cost)

        for row in range(33, 60):
            name = ws_signature.cell(row=row, column=6).value  # Column F
            use = ws_signature.cell(row=row, column=9).value   # Column I
            if name and isinstance(use, (int, float)):
                name = str(name).strip()
                unit_cost = cementing_costs.get(name)
                if unit_cost:
                    cost_records.append({
                        "client": client,
                        "well": well,
                        "basic": basic,
                        "system": system,
                        "materialname": name,
                        "use": use,
                        "cost": unit_cost,
                        "cost * use": use * unit_cost
                    })

        # Step 2: Equipment Tools Cost
        equipment_costs = {}
        for row in ws_price.iter_rows(min_row=16, max_row=36, max_col=8):
            desc = row[1].value or row[2].value
            cost = row[6].value
            if desc and isinstance(cost, (int, float)):
                equipment_costs[str(desc).strip()] = float(cost)

        for row in range(33, 44):
            name = ws_signature.cell(row=row, column=2).value  # Column B
            use = ws_signature.cell(row=row, column=3).value   # Column C
            if name and isinstance(use, (int, float)):
                name = str(name).strip()
                unit_cost = equipment_costs.get(name)
                if unit_cost:
                    cost_records.append({
                        "client": client,
                        "well": well,
                        "basic": basic,
                        "system": system,
                        "materialname": name,
                        "use": use,
                        "cost": unit_cost,
                        "cost * use": use * unit_cost
                    })

        # Step 3: Other Technical Services Cost
        other_costs = {}
        for row in ws_price.iter_rows(min_row=51, max_row=107, max_col=8):
            desc = row[2].value  # Column C
            cost = row[6].value  # Column G
            if desc and isinstance(cost, (int, float)):
                other_costs[str(desc).strip()] = float(cost)

        for row in range(11, 32):
            name = ws_signature.cell(row=row, column=2).value  # Column B
            use = ws_signature.cell(row=row, column=3).value   # Column C
            if name and isinstance(use, (int, float)):
                name = str(name).strip()
                unit_cost = other_costs.get(name)
                if unit_cost:
                    cost_records.append({
                        "client": client,
                        "well": well,
                        "basic": basic,
                        "system": system,
                        "materialname": name,
                        "use": use,
                        "cost": unit_cost,
                        "cost * use": use * unit_cost
                    })

        # Step 4: Casing Accessories Cost
        casing_costs = {}
        for row in ws_price.iter_rows(min_row=111, max_row=209, max_col=8):
            desc = row[2].value  # Column C
            cost = row[6].value  # Column G
            if desc and isinstance(cost, (int, float)):
                casing_costs[str(desc).strip()] = float(cost)

        for row in range(10, 32):
            name = ws_signature.cell(row=row, column=6).value  # Column F
            use = ws_signature.cell(row=row, column=8).value   # Column H
            if name and isinstance(use, (int, float)):
                name = str(name).strip()
                unit_cost = casing_costs.get(name)
                if unit_cost:
                    cost_records.append({
                        "client": client,
                        "well": well,
                        "basic": basic,
                        "system": system,
                        "materialname": name,
                        "use": use,
                        "cost": unit_cost,
                        "cost * use": use * unit_cost
                    })

        # Final Step: Display the cost records
        df_cost_records = pd.DataFrame(cost_records)
        total_cost = df_cost_records["cost * use"].sum()

        # Generating report
        # Revenue records
        revenue_records = {
            'client': [client],
            'well': [well],
            'basic': [basic],
            'system': [system],
            'income': [total_income],
            'well type': [well_type],
            'manager': [manager],
            'area': [area],
            'start time': [start_time],
            'end time': [end_time],
            'well class': [well_class],
            'important type': [important_type],
            'risk': [risk],
            'deep': [deep],
            'tvd': [tvd]
        }

        df_revenue_records = pd.DataFrame(revenue_records)

        # Cost records

        # Profit list
        profit = total_income - total_cost
        profit_rate = (profit/total_income)*100
        profit_list = {
            'client': [client],
            'well': [well],
            'basic': [basic],
            'system': [system],
            'income': [total_income],
            'cost': [total_cost],
            'profit': [profit],
            'profit rate': [profit_rate]
        }

        df_profit_list = pd.DataFrame(profit_list)

    else:
        print('Specify fluid or cement only')
        
    return df_revenue_records, df_cost_records, df_profit_list